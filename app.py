import streamlit as st
import pandas as pd
import io
import json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Sayfa Yapılandırması
st.set_page_config(
    page_title="OOH Planlama Stüdyosu | Medya Yönetim Merkezi",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- ULTRA-PREMIUM EXECUTIVE DYNAMIC CSS ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');

    html, body, .stApp, p, label, input, select, textarea, span, div {
        font-family: 'Outfit', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }

    /* İkon İzolasyonu */
    span[data-testid="stIconMaterial"], .material-symbols-rounded, [class*="material-symbols"] {
        font-family: 'Material Symbols Rounded', 'Material Icons' !important;
    }

    /* Ferah Arka Plan */
    .stApp {
        background: radial-gradient(circle at 50% -10%, #1a294d 0%, #0d1629 50%, #070b14 100%) !important;
        color: #f8fafc !important;
        font-size: 15px !important;
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #101c38 0%, #090e1c 100%) !important;
        border-right: 1px solid rgba(56, 189, 248, 0.15) !important;
    }

    /* Üst Başlık */
    .app-header {
        display: flex;
        align-items: center;
        gap: 14px;
        margin-bottom: 24px;
        padding-bottom: 18px;
        border-bottom: 1px solid rgba(56, 189, 248, 0.2);
    }
    .app-header h1 {
        font-size: 30px !important;
        font-weight: 800 !important;
        letter-spacing: -0.5px;
        background: linear-gradient(135deg, #38bdf8 0%, #a5b4fc 50%, #60a5fa 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin: 0;
    }

    /* Form Etiketleri */
    div[data-testid="stWidgetLabel"] p {
        font-size: 14.5px !important;
        font-weight: 600 !important;
        color: #cbd5e1 !important;
        margin-bottom: 6px !important;
    }

    /* Girdi Kutuları */
    div[data-baseweb="input"], div[data-baseweb="select"] {
        border-radius: 10px !important;
        background-color: #131f3b !important;
        border: 1.5px solid #24355c !important;
        min-height: 48px !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.25) !important;
    }
    div[data-baseweb="input"]:focus-within, div[data-baseweb="select"]:focus-within {
        border-color: #38bdf8 !important;
        box-shadow: 0 0 0 3px rgba(56, 189, 248, 0.3) !important;
    }
    div[data-baseweb="input"] input {
        font-size: 15px !important;
        font-weight: 500 !important;
        color: #ffffff !important;
    }

    /* BUTONLAR - GENEL TABAN */
    .stButton>button, div[data-testid="stPopover"]>button {
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 15.5px !important;
        height: 50px !important;
        padding: 0 24px !important;
        transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
        white-space: nowrap !important;
    }

    /* BUTON ÜZERİNE GELME (HOVER) ETKİSİ */
    .stButton>button:hover, div[data-testid="stPopover"]>button:hover {
        transform: translateY(-3px) scale(1.01) !important;
        cursor: pointer !important;
    }

    /* 1. ÜST SEKME SEÇİM BUTONLARI */
    .tab-btn-active > button {
        background: linear-gradient(135deg, #1d4ed8 0%, #0284c7 100%) !important;
        color: #ffffff !important;
        border: 1.5px solid #38bdf8 !important;
        box-shadow: 0 6px 20px rgba(2, 132, 199, 0.45) !important;
    }
    .tab-btn-active > button:hover {
        box-shadow: 0 8px 25px rgba(56, 189, 248, 0.6) !important;
    }

    .tab-btn-inactive > button {
        background: linear-gradient(135deg, #111a2e 0%, #16223d 100%) !important;
        color: #94a3b8 !important;
        border: 1.5px solid #233354 !important;
    }
    .tab-btn-inactive > button:hover {
        background: linear-gradient(135deg, #1e2c4d 0%, #24355a 100%) !important;
        color: #f8fafc !important;
        border-color: #38bdf8 !important;
        box-shadow: 0 6px 18px rgba(0, 0, 0, 0.4) !important;
    }

    /* 2. EKLE BUTONLARI (Zümrüt Yeşili - Canlı & Ayrışan) */
    .action-add-btn > button {
        background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
        color: #ffffff !important;
        border: 1.5px solid #34d399 !important;
        box-shadow: 0 6px 20px rgba(16, 185, 129, 0.4) !important;
        font-size: 16px !important;
    }
    .action-add-btn > button:hover {
        background: linear-gradient(135deg, #047857 0%, #059669 100%) !important;
        box-shadow: 0 8px 25px rgba(52, 211, 153, 0.6) !important;
        border-color: #6ee7b7 !important;
    }

    /* 3. İKİNCİL / YÖNETİM BUTONLARI */
    button[kind="secondary"], div[data-testid="stPopover"]>button {
        background: linear-gradient(135deg, #1e293b 0%, #131d33 100%) !important;
        color: #e2e8f0 !important;
        border: 1.5px solid #334155 !important;
    }
    button[kind="secondary"]:hover, div[data-testid="stPopover"]>button:hover {
        border-color: #38bdf8 !important;
        box-shadow: 0 6px 18px rgba(56, 189, 248, 0.25) !important;
        color: #ffffff !important;
    }

    /* KPI Metric Kartları */
    div[data-testid="stMetric"] {
        background: linear-gradient(145deg, rgba(26, 38, 68, 0.75) 0%, rgba(15, 23, 42, 0.85) 100%) !important;
        border: 1.5px solid rgba(56, 189, 248, 0.25) !important;
        border-radius: 14px !important;
        padding: 18px 22px !important;
        box-shadow: 0 10px 25px rgba(0, 0, 0, 0.35) !important;
        backdrop-filter: blur(14px) !important;
    }
    div[data-testid="stMetricLabel"] {
        font-size: 13px !important;
        font-weight: 700 !important;
        text-transform: uppercase;
        letter-spacing: 0.9px;
        color: #94a3b8 !important;
    }
    div[data-testid="stMetricValue"] {
        font-size: 30px !important;
        font-weight: 800 !important;
        color: #38bdf8 !important;
        letter-spacing: -0.5px;
    }

    /* Giriş Formu */
    div[data-testid="stForm"] {
        max-width: 480px !important;
        margin: 50px auto 0 auto !important;
        background: linear-gradient(180deg, #15213d 0%, #0e172a 100%) !important;
        border: 1.5px solid rgba(56, 189, 248, 0.3) !important;
        border-radius: 20px !important;
        padding: 38px 32px !important;
        box-shadow: 0 25px 50px rgba(0, 0, 0, 0.7), 0 0 30px rgba(56, 189, 248, 0.15) !important;
    }

    /* Tablo Tasarımı */
    .table-responsive-box {
        width: 100%;
        overflow-x: auto;
        margin: 22px 0 32px 0;
        border: 1.5px solid rgba(56, 189, 248, 0.2);
        border-radius: 14px;
        background-color: #0d1529;
        box-shadow: 0 12px 30px rgba(0,0,0,0.4);
    }
    .custom-ooh-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 14.5px;
    }
    .custom-ooh-table th {
        background: linear-gradient(180deg, #1a2747 0%, #141f38 100%);
        color: #38bdf8;
        padding: 16px 14px;
        text-align: center !important;
        vertical-align: middle;
        font-weight: 700;
        font-size: 13.5px;
        letter-spacing: 0.5px;
        border-bottom: 2px solid #24355a;
        white-space: nowrap;
    }
    .custom-ooh-table td {
        padding: 14px 14px;
        text-align: center !important;
        vertical-align: middle;
        color: #f1f5f9;
        border-bottom: 1px solid #1a2747;
        white-space: nowrap;
    }
    .custom-ooh-table tbody tr:hover {
        background-color: rgba(56, 189, 248, 0.08);
    }

    .corporate-footer {
        text-align: center;
        color: #64748b;
        font-size: 13.5px;
        font-weight: 500;
        margin-top: 55px;
        padding-top: 22px;
        border-top: 1px solid rgba(255, 255, 255, 0.08);
        letter-spacing: 0.5px;
    }
</style>
""", unsafe_allow_html=True)

# --- 1. KULLANICI GİRİŞ SİSTEMİ ---
KULLANICI_ADI = "ibozbek"
KULLANICI_SIFRE = "ibozbek"

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.username = ""

def login_form():
    st.markdown("<div style='height: 40px;'></div>", unsafe_allow_html=True)
    st.markdown("<h2 style='text-align: center; font-weight: 800; font-size: 32px; color: #38bdf8; margin-bottom: 6px; letter-spacing: -0.5px;'>⚡ OOH Planlama Stüdyosu</h2>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #94a3b8; font-size: 15px; margin-bottom: 26px; font-weight: 500;'>Kurumsal Medya Planlama & Simülasyon Portalı</p>", unsafe_allow_html=True)
    
    with st.form("login_box"):
        user = st.text_input("Kullanıcı Adı:", placeholder="Kullanıcı adınızı giriniz")
        pwd = st.text_input("Şifre:", type="password", placeholder="••••••••")
        st.markdown("<div style='height: 14px;'></div>", unsafe_allow_html=True)
        submit = st.form_submit_button("🚀 Güvenli Giriş Yap", use_container_width=True, type="primary")
        if submit:
            if user == KULLANICI_ADI and pwd == KULLANICI_SIFRE:
                st.session_state.logged_in = True
                st.session_state.username = user
                st.rerun()
            else:
                st.error("Kullanıcı adı veya şifre hatalı!")
                
    st.markdown("<div class='corporate-footer'>📌 CAFAS verileri dikkate alınarak geliştirilmiştir.</div>", unsafe_allow_html=True)

if not st.session_state.logged_in:
    login_form()
    st.stop()

# --- 2. SAYI BİÇİMLENDİRME VE ÖZEL İL SAYIMI YARDIMCILARI ---
def tr_tam_sayi(val):
    try:
        n = int(round(float(val)))
        return f"{n:,}".replace(",", ".")
    except:
        return str(val)

def tr_ondalik(val, basamak=2):
    try:
        f = float(val)
        fmt = f"{f:,.{basamak}f}"
        return fmt.replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return str(val)

def format_periyod(val):
    try:
        f_val = float(val)
        if f_val.is_integer():
            return str(int(f_val))
        return tr_ondalik(f_val, 2)
    except:
        return str(val)

def temiz_sayi_al(val, default=0.0):
    if pd.isna(val): return default
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip().replace('%', '')
    if not s or s.lower() in ['nan', 'none']: return default
    if '.' in s and ',' in s:
        if s.rfind(',') > s.rfind('.'): s = s.replace('.', '').replace(',', '.')
        else: s = s.replace(',', '')
    elif ',' in s: s = s.replace(',', '.')
    elif '.' in s:
        parts = s.split('.')
        if len(parts) == 2 and len(parts[1]) == 3 and parts[0].isdigit() and parts[1].isdigit():
            s = parts[0] + parts[1]
    try: return float(s)
    except: return default

def tahmin_mecra(unite_str, il_str):
    u = str(unite_str).lower()
    i = str(il_str).lower()
    if "starbuck" in u or "starbuck" in i:
        return "Core Medya"
    elif "macfit" in u or "mac fit" in u or "macfit" in i or "mac fit" in i:
        return "Donanım Medya"
    elif "üni" in u or "uni" in u or "üniversite" in u:
        return "Üniversite Network"
    elif "istanbul" in i:
        return "İBB / Medya A.Ş."
    elif "ankara" in i or "izmir" in i:
        return "Kentvizyon"
    return "Kentvizyon"

STARBUCKS_ILLERI_SET = {
    "İstanbul", "Ankara", "İzmir", "Bursa", "Antalya", "Adana", "Eskişehir", "Kocaeli", "Gaziantep", 
    "Konya", "Mersin", "Muğla", "Aydın", "Denizli", "Samsun", "Kayseri", "Tekirdağ", "Balıkesir", 
    "Trabzon", "Sakarya", "Çanakkale", "Hatay", "Manisa", "Afyonkarahisar", "Isparta", "Edirne", 
    "Kütahya", "Sivas", "Malatya", "Kahramanmaraş", "Şanlıurfa", "Diyarbakır", "Zonguldak", "Yalova", 
    "Bolu", "Düzce", "Ordu", "Rize"
}

MACFIT_ILLERI_SET = {
    "İstanbul", "Ankara", "İzmir", "Bursa", "Antalya", "Adana", "Eskişehir", "Kocaeli", "Gaziantep", 
    "Konya", "Mersin", "Muğla", "Aydın", "Denizli", "Samsun", "Kayseri", "Tekirdağ", "Balıkesir", 
    "Sakarya", "Yalova"
}

UNI_ILLERI_SET = {
    "İstanbul", "Ankara", "İzmir", "Bursa", "Antalya", "Eskişehir", "Konya", "Trabzon", "Erzurum"
}

def get_il_nufusu(il_adi, nufus_dict):
    il_adi = str(il_adi).strip()
    if "starbuck" in il_adi.lower():
        return float(nufus_dict.get("Starbucks İlleri", 58150000))
    elif "macfit" in il_adi.lower() or "mac fit" in il_adi.lower():
        return float(nufus_dict.get("MacFit İlleri", 52274000))
    elif "üni" in il_adi.lower() or "uni" in il_adi.lower():
        return float(nufus_dict.get("Üni İlleri", 36462000))
    elif "anadolu" in il_adi.lower():
        val = float(nufus_dict.get("Anadolu İlleri", 719000))
        return val if val > 0 else 719000.0
    
    val = float(nufus_dict.get(il_adi, 0))
    if val <= 0:
        val = float(nufus_dict.get("Anadolu İlleri", 719000))
    return val if val > 0 else 719000.0

def hesapla_net_kapsama_metrikleri(df, nufus_dict, tr_total_nufus):
    if df is None or df.empty:
        return 0, 0.0

    kapsanan_tekil_iller = set()
    ozel_ag_nufuslari = []
    
    for _, r in df.iterrows():
        il_str = str(r.get('İl', '')).strip().lower()
        unite_str = str(r.get('Ünite', '')).strip().lower()

        if "starbuck" in il_str or "starbuck" in unite_str:
            kapsanan_tekil_iller.update(STARBUCKS_ILLERI_SET)
            ozel_ag_nufuslari.append(58150000)
        elif "macfit" in il_str or "mac fit" in il_str or "macfit" in unite_str or "mac fit" in unite_str:
            kapsanan_tekil_iller.update(MACFIT_ILLERI_SET)
            ozel_ag_nufuslari.append(52274000)
        elif "üni" in il_str or "uni" in il_str or "üniversite" in il_str or "üni" in unite_str or "uni" in unite_str:
            kapsanan_tekil_iller.update(UNI_ILLERI_SET)
            ozel_ag_nufuslari.append(36462000)
        elif "anadolu" in il_str:
            kapsanan_tekil_iller.add("Anadolu İlleri")
        else:
            kapsanan_tekil_iller.add(r['İl'])

    toplam_il_sayisi = min(81, len(kapsanan_tekil_iller))
    
    net_nufus = 0
    if ozel_ag_nufuslari and len(kapsanan_tekil_iller) <= 38:
        net_nufus = max(ozel_ag_nufuslari)
    else:
        for il in kapsanan_tekil_iller:
            net_nufus += get_il_nufusu(il, nufus_dict)
    
    maks_erisim_pct = min(100.0, round((net_nufus / tr_total_nufus) * 100, 1))
    return toplam_il_sayisi, maks_erisim_pct

# --- 3. EXCEL VERİ MOTORU ---
@st.cache_data
def yerel_exceli_yukle():
    try:
        excel_path = "OUTDOOR.xlsx"
        excel_obj = pd.ExcelFile(excel_path)
        sheet = 'Günlük Gösterim Sayıları' if 'Günlük Gösterim Sayıları' in excel_obj.sheet_names else excel_obj.sheet_names[0]
        df_raw = pd.read_excel(excel_path, sheet_name=sheet, header=None)

        start_row = 1
        for r in range(min(10, len(df_raw))):
            b_val = str(df_raw.iloc[r, 1]).strip().lower() if df_raw.shape[1] > 1 else ""
            c_val = str(df_raw.iloc[r, 2]).strip().lower() if df_raw.shape[1] > 2 else ""
            if "il" in b_val or "ünite" in c_val or "unite" in c_val:
                start_row = r + 1
                break

        rows_data = []
        for r in range(start_row, len(df_raw)):
            il_val = str(df_raw.iloc[r, 1]).strip() if df_raw.shape[1] > 1 else ""
            unite_val = str(df_raw.iloc[r, 2]).strip() if df_raw.shape[1] > 2 else ""

            if not il_val or il_val.lower() in ['nan', 'none', '', 'il', 'i̇l'] or not unite_val or unite_val.lower() in ['nan', 'none', '', 'ünite', 'unite']:
                continue

            gost_val = temiz_sayi_al(df_raw.iloc[r, 3] if df_raw.shape[1] > 3 else 0, 0.0)
            frek_val = temiz_sayi_al(df_raw.iloc[r, 4] if df_raw.shape[1] > 4 else 1, 1.0)
            net_val = temiz_sayi_al(df_raw.iloc[r, 5] if df_raw.shape[1] > 5 else 100, 100.0)
            endeks_raw = temiz_sayi_al(df_raw.iloc[r, 6] if df_raw.shape[1] > 6 else 1, 1.0)
            endeks_val = endeks_raw / 100.0 if endeks_raw > 1.5 else endeks_raw

            rows_data.append({
                'İl': il_val,
                'Ünite': unite_val,
                'Günlük Gösterim': gost_val,
                'Frekans': frek_val,
                'Network Adedi': int(round(net_val)),
                'Endeks': endeks_val
            })

        df_gost = pd.DataFrame(rows_data)

        nufus_dict = {
            "İstanbul": 15754053, "Ankara": 5910320, "İzmir": 4504185, 
            "Bursa": 3263011, "Antalya": 2777677, "Adana": 2274106, "Konya": 2320645,
            "Gaziantep": 2185982, "Kocaeli": 2102907, "Mersin": 1938389, "Diyarbakır": 1818133,
            "Hatay": 1544640, "Manisa": 1475716, "Kayseri": 1445495, "Samsun": 1377546,
            "Balıkesir": 1273519, "Tekirdağ": 1167059, "Aydın": 1161702, "Van": 1127612,
            "Trabzon": 824352, "Eskişehir": 915418, "Denizli": 1059082, "Sakarya": 1098115,
            "Muğla": 1066736, "Türkiye": 86920168, "Starbucks İlleri": 58150000,
            "MacFit İlleri": 52274000, "Üni İlleri": 36462000, "Anadolu İlleri": 719000
        }
        sure_dict = {
            "Durak Raket CLP": 7, "Billboard": 7, "Afiş Değiştiricili Megalight": 7,
            "Megalight": 7, "Üst Geçit Alınlık": 15, "Giantboard": 10,
            "Elektrik Direği Banner": 14, "Avm Dış Led Ekran": 7, "Dijital Raket": 7,
            "Dijital Ekran": 7, "Toplu Taşıma Ekran": 7, "Tramvay Raket CLP": 7, "Raket CLP": 7,
            "Üni Ekran": 7, "Üniversite Ekran": 7, "Starbucks Kasa Arkası Ekran": 7,
            "MacFit Ekran": 7
        }

        for c in range(5, df_raw.shape[1]):
            col_txt = " ".join([str(v) for v in df_raw.iloc[:, c].dropna().values])
            if "Nüfus" in col_txt:
                for r in range(start_row - 1, len(df_raw)):
                    il_c = str(df_raw.iloc[r, c-1] if c>=1 else "").strip()
                    nuf_val = temiz_sayi_al(df_raw.iloc[r, c], None)
                    if il_c and nuf_val is not None and il_c.lower() != 'nan' and nuf_val > 0:
                        nufus_dict[il_c] = nuf_val

            if "Kullanım Süresi" in col_txt or "Süre" in col_txt:
                for r in range(start_row - 1, len(df_raw)):
                    u_c = str(df_raw.iloc[r, c-2] if c>=2 else df_raw.iloc[r, c-1]).strip()
                    sure_val = temiz_sayi_al(df_raw.iloc[r, c], None)
                    if u_c and sure_val is not None and u_c.lower() != 'nan' and sure_val > 0:
                        sure_dict[u_c] = sure_val

        tr_nufus = float(nufus_dict.get("Türkiye", 86920168))
        if nufus_dict.get("Anadolu İlleri", 0) <= 0:
            nufus_dict["Anadolu İlleri"] = 719000.0

        return df_gost, nufus_dict, sure_dict, tr_nufus
    except Exception as e:
        st.error(f"Excel Okuma Hatası (OUTDOOR.xlsx): {e}")
        return None, {}, {}, 86920168

df_gost, nufus_dict, sure_dict, TR_TOTAL_NUFUS = yerel_exceli_yukle()

# --- 4. SESSION STATE ---
if "active_tab" not in st.session_state:
    st.session_state.active_tab = "simulasyon"

if "sim_rows" not in st.session_state:
    st.session_state.sim_rows = []

if "arsiv_rows" not in st.session_state:
    st.session_state.arsiv_rows = []

if "sim_per" not in st.session_state:
    st.session_state.sim_per = 1.0
if "sim_sure" not in st.session_state:
    st.session_state.sim_sure = 7

if "ars_per" not in st.session_state:
    st.session_state.ars_per = 1.0
if "ars_sure" not in st.session_state:
    st.session_state.ars_sure = 7

# --- 5. YAN PANEL (SIDEBAR) ---
st.sidebar.markdown(f"**👤 Giriş Yapan:** `{st.session_state.username}`")
if st.sidebar.button("🚪 Çıkış Yap"):
    st.session_state.logged_in = False
    st.session_state.username = ""
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.subheader("⚙️ Sistem Ayarları")

if st.sidebar.button("🔄 Excel Verisini Yenile", use_container_width=True):
    st.cache_data.clear()
    st.rerun()

looker_url = st.sidebar.text_input(
    "🗺️ Looker Studio Harita Linki:",
    placeholder="https://lookerstudio.google.com/embed/reporting/..."
)

st.sidebar.markdown("---")
st.sidebar.caption("⚡ **Geliştirici:** İbrahim Özbek Arslan")

# --- 6. RAPOR OLUŞTURMA YARDIMCILARI (EXCEL & HTML) ---
def generate_excel_report(df_to_export, report_title, looker_link="", is_arsiv=False):
    output = io.BytesIO()
    df_excel = df_to_export.copy()
    if "TR Erişim %" in df_excel.columns:
        df_excel["TR Erişim %"] = df_excel["TR Erişim %"] / 100.0

    toplam_gos = int(df_to_export["Toplam Gösterim"].sum())
    toplam_grp = float(round(df_to_export["TR GRP"].sum(), 2))
    kapsanan_il, maks_erisim = hesapla_net_kapsama_metrikleri(df_to_export, nufus_dict, TR_TOTAL_NUFUS)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df = pd.DataFrame([
            {"Metrik": "Rapor Başlığı", "Değer": report_title},
            {"Metrik": "Geliştirici & Sistem", "Değer": "İbrahim Özbek Arslan | OOH Planlama Stüdyosu"},
            {"Metrik": "Toplam Gösterim", "Değer": toplam_gos},
            {"Metrik": "Toplam TR GRP", "Değer": toplam_grp},
            {"Metrik": "Kapsanan İl Sayısı", "Değer": f"{kapsanan_il} İl"},
            {"Metrik": "Maks. TR Erişimi", "Değer": float(maks_erisim / 100.0)},
            {"Metrik": "Looker Studio Harita Linki", "Değer": looker_link if looker_link else "Belirtilmedi"}
        ])
        
        summary_df.to_excel(writer, sheet_name='Özet KPI', index=False)
        df_excel.to_excel(writer, sheet_name='Medya Planı', index=False)
        
        wb = writer.book
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="38BDF8")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws_sum = wb['Özet KPI']
        for col in ws_sum.columns:
            for cell in col:
                cell.border = thin_border
                cell.alignment = left_align
        for cell in ws_sum[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        ws_sum["B3"].number_format = '#,##0'
        ws_sum["B4"].number_format = '#,##0.00'
        ws_sum["B6"].number_format = '0.0%'
        
        ws_plan = wb['Medya Planı']
        for col in ws_plan.columns:
            for cell in col:
                cell.border = thin_border
                cell.alignment = center_align
                
        for cell in ws_plan[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        col_names = [cell.value for cell in ws_plan[1]]
        for row in range(2, ws_plan.max_row + 1):
            for col_idx, col_name in enumerate(col_names, start=1):
                cell = ws_plan.cell(row=row, column=col_idx)
                if col_name in ["Adet", "Toplam Gösterim", "Erişim (Kişi)", "İl Nüfusu", "TR Nüfusu"]:
                    cell.number_format = '#,##0'
                elif col_name in ["Frekans", "TR GRP"]:
                    cell.number_format = '#,##0.00'
                elif col_name in ["TR Erişim %"]:
                    cell.number_format = '0.00%'

        for ws in [ws_sum, ws_plan]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    return output.getvalue()

def generate_html_report(df_to_export, report_title, include_looker=False, is_arsiv=False):
    if is_arsiv:
        table_headers = "<th>Yıl</th><th>Dönem</th><th>Marka</th><th>Kampanya</th><th>Mecra</th><th>Ünite</th><th>İl</th><th>Süre (Gün)</th><th>Periyod</th><th>Adet</th><th>Toplam Gösterim</th><th>Frekans</th><th>Erişim (Kişi)</th><th>İl Nüfusu</th><th>TR Nüfusu</th><th>TR Erişim %</th><th>TR GRP</th>"
        rows_list = []
        for _, r in df_to_export.iterrows():
            rows_list.append(f"<tr><td>{r['Yıl']}</td><td>{r['Dönem (Ay)']}</td><td>{r['Marka']}</td><td>{r['Kampanya Adı']}</td><td>{r['Mecra Adı']}</td><td>{r['Ünite']}</td><td>{r['İl']}</td><td>{r['Süre (Gün)']}</td><td>{r['Periyod']}</td><td>{tr_tam_sayi(r['Adet'])}</td><td>{tr_tam_sayi(r['Toplam Gösterim'])}</td><td>{tr_ondalik(r['Frekans'], 1)}</td><td>{tr_tam_sayi(r['Erişim (Kişi)'])}</td><td>{tr_tam_sayi(r['İl Nüfusu'])}</td><td>{tr_tam_sayi(r['TR Nüfusu'])}</td><td>%{tr_ondalik(r['TR Erişim %'], 2)}</td><td>{tr_ondalik(r['TR GRP'], 2)}</td></tr>")
        table_rows_html = "".join(rows_list)
    else:
        table_headers = "<th>Ünite</th><th>İl</th><th>Süre (Gün)</th><th>Periyod</th><th>Adet</th><th>Toplam Gösterim</th><th>Frekans</th><th>Erişim (Kişi)</th><th>İl Nüfusu</th><th>TR Nüfusu</th><th>TR Erişim %</th><th>TR GRP</th>"
        rows_list = []
        for _, r in df_to_export.iterrows():
            rows_list.append(f"<tr><td>{r['Ünite']}</td><td>{r['İl']}</td><td>{r['Süre (Gün)']}</td><td>{r['Periyod']}</td><td>{tr_tam_sayi(r['Adet'])}</td><td>{tr_tam_sayi(r['Toplam Gösterim'])}</td><td>{tr_ondalik(r['Frekans'], 1)}</td><td>{tr_tam_sayi(r['Erişim (Kişi)'])}</td><td>{tr_tam_sayi(r['İl Nüfusu'])}</td><td>{tr_tam_sayi(r['TR Nüfusu'])}</td><td>%{tr_ondalik(r['TR Erişim %'], 2)}</td><td>{tr_ondalik(r['TR GRP'], 2)}</td></tr>")
        table_rows_html = "".join(rows_list)

    toplam_gos = df_to_export["Toplam Gösterim"].sum()
    toplam_grp = round(df_to_export["TR GRP"].sum(), 2)
    kapsanan_il, maks_erisim = hesapla_net_kapsama_metrikleri(df_to_export, nufus_dict, TR_TOTAL_NUFUS)

    looker_section = ""
    if include_looker and looker_url:
        looker_section = f"""<div style="margin-top: 30px; background: #131b2e; border: 1px solid #1f293d; border-radius: 10px; padding: 20px;"><h3 style="color: #38bdf8; margin-bottom: 15px;">🗺️ Kampanya Harita ve Lokasyon Paneli</h3><iframe src="{looker_url}" width="100%" height="560" frameborder="0" style="border:0; border-radius: 8px;" allowfullscreen></iframe></div>"""

    return f"""<!DOCTYPE html>
<html lang="tr">
<head>
    <meta charset="UTF-8">
    <title>{report_title}</title>
    <style>
        body {{ background-color: #090d16; color: #e2e8f0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; padding: 30px; line-height: 1.5; }}
        .kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 25px; }}
        .kpi-card {{ background: #131b2e; border: 1px solid #1f293d; border-radius: 10px; padding: 20px; text-align: center; }}
        table {{ width: 100%; border-collapse: collapse; text-align: center; font-size: 13px; margin-top: 20px; }}
        th {{ background: #1e293b; color: #38bdf8; padding: 12px; border: 1px solid #1f293d; text-align: center; }}
        td {{ padding: 10px; border: 1px solid #1f293d; color: #cbd5e1; text-align: center; }}
        tr:nth-child(even) {{ background: rgba(255,255,255,0.02); }}
        .footer-note {{ text-align: center; color: #94a3b8; font-size: 13px; margin-top: 40px; padding-top: 20px; border-top: 1px solid #1e293b; }}
    </style>
</head>
<body>
    <h1 style="color: #f1f5f9;">📊 {report_title}</h1>
    <div class="kpi-grid">
        <div class="kpi-card"><div style="font-size: 12px; color: #94a3b8;">TOPLAM GÖSTERİM</div><div style="font-size: 24px; font-weight: bold; color: #4ade80;">{tr_tam_sayi(toplam_gos)}</div></div>
        <div class="kpi-card"><div style="font-size: 12px; color: #94a3b8;">TOPLAM TR GRP</div><div style="font-size: 24px; font-weight: bold; color: #38bdf8;">{tr_ondalik(toplam_grp, 2)}</div></div>
        <div class="kpi-card"><div style="font-size: 12px; color: #94a3b8;">KAPSASANAN İL</div><div style="font-size: 24px; font-weight: bold; color: #c084fc;">{kapsanan_il} İl</div></div>
        <div class="kpi-card"><div style="font-size: 12px; color: #94a3b8;">MAKS. TR ERİŞİMİ</div><div style="font-size: 24px; font-weight: bold; color: #facc15;">%{tr_ondalik(maks_erisim, 1)}</div></div>
    </div>
    <table>
        <thead><tr>{table_headers}</tr></thead>
        <tbody>{table_rows_html}</tbody>
    </table>
    {looker_section}
    <div class="footer-note">📌 CAFAS verileri dikkate alınarak geliştirilmiştir.</div>
</body>
</html>"""

# --- 7. ÜST MENÜ & BAŞLIK ---
st.markdown("""
<div class="app-header">
    <h1>⚡ OOH PLANLAMA & SİMÜLASYON MERKEZİ</h1>
</div>
""", unsafe_allow_html=True)

# SEKME GEÇİŞ BUTONLARI (Aktif ve Pasif Sınıflarıyla Ayrıştırıldı)
col_btn1, col_btn2 = st.columns(2)
with col_btn1:
    btn1_class = "tab-btn-active" if st.session_state.active_tab == "simulasyon" else "tab-btn-inactive"
    st.markdown(f'<div class="{btn1_class}">', unsafe_allow_html=True)
    if st.button("📊 Anlık Hesaplama & Simülatör", key="tab_sim_btn", use_container_width=True):
        st.session_state.active_tab = "simulasyon"
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

with col_btn2:
    btn2_class = "tab-btn-active" if st.session_state.active_tab == "arsiv" else "tab-btn-inactive"
    st.markdown(f'<div class="{btn2_class}">', unsafe_allow_html=True)
    if st.button("📁 Kampanya Yönetimi & Yıllık Arşiv", key="tab_ars_btn", use_container_width=True):
        st.session_state.active_tab = "arsiv"
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown("<div style='height: 15px;'></div>", unsafe_allow_html=True)

# ==========================================
# 1. SEKME: ANLIK HESAPLAMA & SİMÜLATÖR
# ==========================================
if st.session_state.active_tab == "simulasyon":
    if df_gost is not None and not df_gost.empty:
        st.markdown("<h4 style='color: #94a3b8; font-weight: 700; font-size: 16px; margin-bottom: 12px;'>📝 YENİ KAMPANYA SİMÜLASYONU</h4>", unsafe_allow_html=True)

        col_il, col_unite, col_per, col_sure, col_adet = st.columns([2.2, 2.2, 1.2, 1.2, 1.2])
        
        with col_il:
            il_listesi = sorted(list(set(df_gost['İl'].tolist())))
            
            def on_sim_il_change():
                sec_il = st.session_state.sim_il_select
                uniteler_yeni = sorted(list(set(df_gost[df_gost['İl'] == sec_il]['Ünite'].tolist())))
                if uniteler_yeni:
                    ilk_unite = uniteler_yeni[0]
                    b = sure_dict.get(ilk_unite, 7.0)
                    st.session_state.sim_sure = int(round(b * st.session_state.sim_per))
                    row_match = df_gost[(df_gost['İl'] == sec_il) & (df_gost['Ünite'] == ilk_unite)]
                    if not row_match.empty:
                        st.session_state.sim_adet_input = int(row_match['Network Adedi'].values[0])

            secilen_il = st.selectbox("📍 İl Seçin:", il_listesi, key="sim_il_select", on_change=on_sim_il_change)

        with col_unite:
            uniteler = sorted(list(set(df_gost[df_gost['İl'] == secilen_il]['Ünite'].tolist())))
            
            def on_sim_unite_change():
                u = st.session_state.sim_unite_select
                b = sure_dict.get(u, 7.0)
                st.session_state.sim_sure = int(round(b * st.session_state.sim_per))
                
                row_match = df_gost[(df_gost['İl'] == st.session_state.sim_il_select) & (df_gost['Ünite'] == u)]
                if not row_match.empty:
                    st.session_state.sim_adet_input = int(row_match['Network Adedi'].values[0])

            secilen_unite = st.selectbox("🎯 Ünite Seçin:", uniteler, key="sim_unite_select", on_change=on_sim_unite_change)
            baz_sure = sure_dict.get(secilen_unite, 7.0)

            current_net_row = df_gost[(df_gost['İl'] == secilen_il) & (df_gost['Ünite'] == secilen_unite)]
            current_net_adet = int(current_net_row['Network Adedi'].values[0]) if not current_net_row.empty else 100

        def update_from_per():
            p = st.session_state.sim_per
            st.session_state.sim_sure = int(round(baz_sure * p))

        def update_from_sure():
            s = st.session_state.sim_sure
            st.session_state.sim_per = round(s / baz_sure, 2) if baz_sure > 0 else 1.0

        with col_per:
            st.number_input(
                "⏱️ Periyod:",
                min_value=0.1,
                max_value=20.0,
                step=0.1,
                key="sim_per",
                on_change=update_from_per
            )

        with col_sure:
            st.number_input(
                "📅 Süre (Gün):",
                min_value=1,
                step=1,
                key="sim_sure",
                on_change=update_from_sure
            )

        with col_adet:
            if "sim_adet_input" not in st.session_state:
                st.session_state.sim_adet_input = current_net_adet
            adet_val = st.number_input("🔢 Adet:", min_value=1, value=int(st.session_state.sim_adet_input), step=1, key="sim_adet_input")

        periyod_val = st.session_state.sim_per
        sure_val = st.session_state.sim_sure

        # ÖZEL ZÜMRÜT YEŞİLİ EKLE BUTONU
        st.markdown('<div class="action-add-btn">', unsafe_allow_html=True)
        ekle_tiklandi = st.button("➕ Simülasyon Satırını Plana Ekle", key="sim_add_row_btn", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        if ekle_tiklandi:
            m_gost = df_gost[(df_gost['İl'] == secilen_il) & (df_gost['Ünite'] == secilen_unite)]
            gunluk_gost = float(m_gost['Günlük Gösterim'].values[0]) if not m_gost.empty else 0.0
            baz_frekans = float(m_gost['Frekans'].values[0]) if not m_gost.empty else 1.0
            network_adedi = float(m_gost['Network Adedi'].values[0]) if not m_gost.empty else 100.0
            endeks = float(m_gost['Endeks'].values[0]) if not m_gost.empty else 1.0

            if network_adedi > 0 and adet_val > 0 and periyod_val > 0:
                dinamik_frekans = baz_frekans * ((adet_val / network_adedi) ** 0.55) * endeks * (periyod_val ** 0.80)
            else:
                dinamik_frekans = 0.0

            il_nufus = get_il_nufusu(secilen_il, nufus_dict)
            toplam_gosterim = gunluk_gost * sure_val * adet_val
            erisim_kisi = (toplam_gosterim / dinamik_frekans) if dinamik_frekans > 0 else 0
            erisim_pct_tr = (erisim_kisi / TR_TOTAL_NUFUS) * 100
            grp_tr = (toplam_gosterim / TR_TOTAL_NUFUS) * 100

            st.session_state.sim_rows.append({
                "Ünite": secilen_unite,
                "İl": secilen_il,
                "Süre (Gün)": int(sure_val),
                "Periyod": format_periyod(periyod_val),
                "Adet": int(adet_val),
                "Toplam Gösterim": int(toplam_gosterim),
                "Frekans": float(round(dinamik_frekans, 1)),
                "Erişim (Kişi)": int(erisim_kisi),
                "İl Nüfusu": int(il_nufus),
                "TR Nüfusu": int(TR_TOTAL_NUFUS),
                "TR Erişim %": float(round(erisim_pct_tr, 2)),
                "TR GRP": float(round(grp_tr, 2))
            })
            st.rerun()

        if st.session_state.sim_rows:
            df_sim = pd.DataFrame(st.session_state.sim_rows)
            st.markdown("<div style='height: 18px;'></div>", unsafe_allow_html=True)
            kpi1, kpi2, kpi3, kpi4 = st.columns(4)
            toplam_gos = df_sim["Toplam Gösterim"].sum()
            toplam_grp = round(df_sim["TR GRP"].sum(), 2)
            kapsanan_il, maks_erisim = hesapla_net_kapsama_metrikleri(df_sim, nufus_dict, TR_TOTAL_NUFUS)

            kpi1.metric("📊 Toplam Gösterim", tr_tam_sayi(toplam_gos))
            kpi2.metric("🇹🇷 Toplam TR GRP", tr_ondalik(toplam_grp, 2))
            kpi3.metric("🌐 Maks. TR Erişimi", f"%{tr_ondalik(maks_erisim, 1)}")
            kpi4.metric("📍 Kapsanan İl Sayısı", f"{kapsanan_il} İl")

            rows_html = "".join([
                f"<tr><td>{r['Ünite']}</td><td>{r['İl']}</td><td>{r['Süre (Gün)']}</td><td>{r['Periyod']}</td><td>{tr_tam_sayi(r['Adet'])}</td><td>{tr_tam_sayi(r['Toplam Gösterim'])}</td><td>{tr_ondalik(r['Frekans'], 1)}</td><td>{tr_tam_sayi(r['Erişim (Kişi)'])}</td><td>{tr_tam_sayi(r['İl Nüfusu'])}</td><td>{tr_tam_sayi(r['TR Nüfusu'])}</td><td>%{tr_ondalik(r['TR Erişim %'], 2)}</td><td>{tr_ondalik(r['TR GRP'], 2)}</td></tr>"
                for _, r in df_sim.iterrows()
            ])
            
            table_markup = f"""<div class="table-responsive-box"><table class="custom-ooh-table"><thead><tr><th>Ünite</th><th>İl</th><th>Süre (Gün)</th><th>Periyod</th><th>Adet</th><th>Toplam Gösterim</th><th>Frekans</th><th>Erişim (Kişi)</th><th>İl Nüfusu</th><th>TR Nüfusu</th><th>TR Erişim %</th><th>TR GRP</th></tr></thead><tbody>{rows_html}</tbody></table></div>"""
            st.markdown(table_markup, unsafe_allow_html=True)

            col_s1, col_s2, col_s3, col_s4, col_s5, col_s6 = st.columns([1, 1.2, 1, 1.2, 1.2, 1.5])
            with col_s1:
                if st.button("↩️ Son Satırı Sil", key="sim_del_last", use_container_width=True):
                    if st.session_state.sim_rows:
                        st.session_state.sim_rows.pop()
                        st.rerun()
            with col_s2:
                with st.popover("🗑️ Seçili Satırı Sil", use_container_width=True):
                    silinecek_sim_idx = st.selectbox(
                        "Silinecek Satır No:",
                        range(len(st.session_state.sim_rows)),
                        key="sim_del_select",
                        format_func=lambda i: f"Satır {i+1}: {st.session_state.sim_rows[i]['Ünite']} ({st.session_state.sim_rows[i]['İl']})"
                    )
                    if st.button("❌ Bu Satırı Sil", key="sim_del_btn", type="primary", use_container_width=True):
                        st.session_state.sim_rows.pop(silinecek_sim_idx)
                        st.rerun()
            with col_s3:
                if st.button("🧹 Tümünü Temizle", key="sim_clear_all", use_container_width=True):
                    st.session_state.sim_rows = []
                    st.rerun()
            with col_s4:
                sim_html = generate_html_report(df_sim, "OOH Medya Simülasyon Raporu", include_looker=True, is_arsiv=False)
                st.download_button(
                    label="📄 HTML Raporu Al",
                    data=sim_html,
                    file_name="OOH_Simulasyon_Raporu.html",
                    mime="text/html",
                    use_container_width=True
                )
            with col_s5:
                sim_excel = generate_excel_report(df_sim, "OOH Medya Simülasyon Raporu", looker_link=looker_url, is_arsiv=False)
                st.download_button(
                    label="📊 Excel Raporu Al",
                    data=sim_excel,
                    file_name="OOH_Simulasyon_Raporu.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with col_s6:
                with st.popover("📁 Arşive Aktar", use_container_width=True):
                    st.markdown("##### 📌 Genel Kampanya Bilgileri")
                    aktar_yil = st.number_input("Yıl:", min_value=2020, max_value=2035, value=2026, step=1, key="aktar_yil")
                    aylar = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
                    aktar_donem = st.selectbox("Dönem:", aylar, index=0, key="aktar_donem")
                    aktar_marka = st.text_input("Marka:", placeholder="Örn: BİM", key="aktar_marka")
                    aktar_kampanya = st.text_input("Kampanya Adı:", placeholder="Örn: Kırtasiye", key="aktar_kampanya")

                    st.markdown("---")
                    st.markdown("##### 🏢 Satır Bazlı Mecra Eşleştirmesi")
                    
                    mecra_girdileri = []
                    for idx, row in enumerate(st.session_state.sim_rows):
                        varsayilan_mecra = tahmin_mecra(row['Ünite'], row['İl'])
                        m_val = st.text_input(
                            f"{row['Ünite']} ({row['İl']}):",
                            value=varsayilan_mecra,
                            key=f"mecra_input_{idx}"
                        )
                        mecra_girdileri.append(m_val.strip() if m_val.strip() else varsayilan_mecra)

                    if st.button("✅ Arşive Gönder", use_container_width=True, type="primary"):
                        m_isim = aktar_marka.strip() if aktar_marka.strip() else "BİM"
                        k_isim = aktar_kampanya.strip() if aktar_kampanya.strip() else "Kırtasiye"

                        for idx, row in enumerate(st.session_state.sim_rows):
                            c_isim = mecra_girdileri[idx]
                            st.session_state.arsiv_rows.append({
                                "Yıl": int(aktar_yil),
                                "Dönem (Ay)": aktar_donem,
                                "Marka": m_isim,
                                "Kampanya Adı": k_isim,
                                "Mecra Adı": c_isim,
                                "Ünite": row["Ünite"],
                                "İl": row["İl"],
                                "Süre (Gün)": row["Süre (Gün)"],
                                "Periyod": row["Periyod"],
                                "Adet": int(row["Adet"]),
                                "Toplam Gösterim": int(row["Toplam Gösterim"]),
                                "Frekans": float(row["Frekans"]),
                                "Erişim (Kişi)": int(row["Erişim (Kişi)"]),
                                "İl Nüfusu": int(row["İl Nüfusu"]),
                                "TR Nüfusu": int(row["TR Nüfusu"]),
                                "TR Erişim %": float(row["TR Erişim %"]),
                                "TR GRP": float(row["TR GRP"])
                            })
                        st.success(f"✅ {len(st.session_state.sim_rows)} satır kendi mecralarıyla arşive aktarıldı!")
                        st.rerun()

        # Harita Paneli
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        st.markdown("<h4 style='color: #94a3b8; font-weight: 700; font-size: 16px; margin-bottom: 12px;'>🗺️ CANLI LOOKER STUDIO HARİTA PANELİ</h4>", unsafe_allow_html=True)
        if looker_url:
            st.components.v1.html(
                f'<iframe src="{looker_url}" width="100%" height="540" frameborder="0" style="border:0; border-radius: 12px; box-shadow: 0 10px 30px rgba(0,0,0,0.4);" allowfullscreen></iframe>',
                height=560
            )
        else:
            st.info("💡 Haritayı görüntülemek için sol yan menüden **Looker Studio Harita Linki**ni giriniz.")

# ==========================================
# 2. SEKME: KAMPANYA YÖNETİMİ & YILLIK ARŞİV
# ==========================================
elif st.session_state.active_tab == "arsiv":
    st.markdown("<h4 style='color: #94a3b8; font-weight: 700; font-size: 16px; margin-bottom: 12px;'>📝 YENİ KAMPANYA SATIRI EKLE</h4>", unsafe_allow_html=True)
    
    if df_gost is not None and not df_gost.empty:
        il_listesi = sorted(list(set(df_gost['İl'].tolist())))

        # 1. Satır: Yıl, Dönem, Marka, Kampanya, Mecra
        k1, k2, k3, k4, k5 = st.columns([1.2, 1.3, 2.5, 2.5, 2.5])
        with k1:
            a_yil = st.number_input("Yıl:", min_value=2020, max_value=2035, value=2026, step=1, key="ars_yil")
        with k2:
            aylar = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
            a_donem = st.selectbox("Dönem:", aylar, index=0, key="ars_donem")
        with k3:
            a_marka_in = st.text_input("Marka:", placeholder="Örn: BİM", key="ars_marka")
            a_marka = a_marka_in.strip() if a_marka_in.strip() else "BİM"
        with k4:
            a_kampanya_in = st.text_input("Kampanya:", placeholder="Örn: Kırtasiye", key="ars_kampanya")
            a_kampanya = a_kampanya_in.strip() if a_kampanya_in.strip() else "Kırtasiye"
        with k5:
            a_mecra_in = st.text_input("Mecra:", placeholder="Örn: Kentvizyon / Donanım Medya", key="ars_mecra")
            a_mecra = a_mecra_in.strip() if a_mecra_in.strip() else "Kentvizyon"

        # 2. Satır: İl, Ünite, Periyod, Süre, Adet, Ekle
        k6, k7, k8, k9, k10, k11 = st.columns([2.2, 2.5, 1.2, 1.2, 1.2, 1.7])
        with k6:
            def on_ars_il_change():
                sec_il_a = st.session_state.ars_il_select
                uniteler_yeni_a = sorted(list(set(df_gost[df_gost['İl'] == sec_il_a]['Ünite'].tolist())))
                if uniteler_yeni_a:
                    ilk_u = uniteler_yeni_a[0]
                    b = sure_dict.get(ilk_u, 7.0)
                    st.session_state.ars_sure = int(round(b * st.session_state.ars_per))
                    row_match_a = df_gost[(df_gost['İl'] == sec_il_a) & (df_gost['Ünite'] == ilk_u)]
                    if not row_match_a.empty:
                        st.session_state.ars_adet_input = int(row_match_a['Network Adedi'].values[0])

            a_il = st.selectbox("İl:", il_listesi, key="ars_il_select", on_change=on_ars_il_change)

        with k7:
            a_uniteler = sorted(list(set(df_gost[df_gost['İl'] == a_il]['Ünite'].tolist())))
            
            def on_ars_unite_change():
                u = st.session_state.ars_unite_select
                b = sure_dict.get(u, 7.0)
                st.session_state.ars_sure = int(round(b * st.session_state.ars_per))
                
                row_match_a = df_gost[(df_gost['İl'] == st.session_state.ars_il_select) & (df_gost['Ünite'] == u)]
                if not row_match_a.empty:
                    st.session_state.ars_adet_input = int(row_match_a['Network Adedi'].values[0])

            a_unite = st.selectbox("Ünite:", a_uniteler, key="ars_unite_select", on_change=on_ars_unite_change)
            baz_sure_a = sure_dict.get(a_unite, 7.0)
            
            current_net_row_a = df_gost[(df_gost['İl'] == a_il) & (df_gost['Ünite'] == a_unite)]
            current_net_adet_a = int(current_net_row_a['Network Adedi'].values[0]) if not current_net_row_a.empty else 100

        def update_from_ars_per():
            p = st.session_state.ars_per
            st.session_state.ars_sure = int(round(baz_sure_a * p))

        def update_from_ars_sure():
            s = st.session_state.ars_sure
            st.session_state.ars_per = round(s / baz_sure_a, 2) if baz_sure_a > 0 else 1.0

        with k8:
            st.number_input(
                "Periyod:",
                min_value=0.1,
                max_value=20.0,
                step=0.1,
                key="ars_per",
                on_change=update_from_ars_per
            )
        with k9:
            st.number_input(
                "Süre:",
                min_value=1,
                step=1,
                key="ars_sure",
                on_change=update_from_ars_sure
            )
        with k10:
            if "ars_adet_input" not in st.session_state:
                st.session_state.ars_adet_input = current_net_adet_a
            a_adet = st.number_input("Adet:", min_value=1, value=int(st.session_state.ars_adet_input), step=1, key="ars_adet_input")
        with k11:
            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
            # ÖZEL ZÜMRÜT YEŞİLİ EKLE BUTONU
            st.markdown('<div class="action-add-btn">', unsafe_allow_html=True)
            ekle_btn = st.button("➕ Ekle", key="ars_add_btn", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        a_periyod = st.session_state.ars_per
        a_sure = st.session_state.ars_sure

        if ekle_btn:
            m_gost = df_gost[(df_gost['İl'] == a_il) & (df_gost['Ünite'] == a_unite)]
            gunluk_gost = float(m_gost['Günlük Gösterim'].values[0]) if not m_gost.empty else 0.0
            baz_frekans = float(m_gost['Frekans'].values[0]) if not m_gost.empty else 1.0
            network_adedi = float(m_gost['Network Adedi'].values[0]) if not m_gost.empty else 100.0
            endeks = float(m_gost['Endeks'].values[0]) if not m_gost.empty else 1.0

            if network_adedi > 0 and a_adet > 0 and a_periyod > 0:
                dinamik_frekans = baz_frekans * ((a_adet / network_adedi) ** 0.55) * endeks * (a_periyod ** 0.80)
            else:
                dinamik_frekans = 0.0

            il_nufus = get_il_nufusu(a_il, nufus_dict)
            toplam_gosterim = gunluk_gost * a_sure * a_adet
            erisim_kisi = (toplam_gosterim / dinamik_frekans) if dinamik_frekans > 0 else 0
            erisim_pct_tr = (erisim_kisi / TR_TOTAL_NUFUS) * 100
            grp_tr = (toplam_gosterim / TR_TOTAL_NUFUS) * 100

            st.session_state.arsiv_rows.append({
                "Yıl": int(a_yil),
                "Dönem (Ay)": a_donem,
                "Marka": a_marka,
                "Kampanya Adı": a_kampanya,
                "Mecra Adı": a_mecra,
                "Ünite": a_unite,
                "İl": a_il,
                "Süre (Gün)": int(a_sure),
                "Periyod": format_periyod(a_periyod),
                "Adet": int(a_adet),
                "Toplam Gösterim": int(toplam_gosterim),
                "Frekans": float(round(dinamik_frekans, 1)),
                "Erişim (Kişi)": int(erisim_kisi),
                "İl Nüfusu": int(il_nufus),
                "TR Nüfusu": int(TR_TOTAL_NUFUS),
                "TR Erişim %": float(round(erisim_pct_tr, 2)),
                "TR GRP": float(round(grp_tr, 2))
            })
            st.rerun()

        if st.session_state.arsiv_rows:
            df_arsiv = pd.DataFrame(st.session_state.arsiv_rows)
            st.markdown("<div style='height: 18px;'></div>", unsafe_allow_html=True)
            
            ak1, ak2, ak3, ak4 = st.columns(4)
            toplam_gos_a = df_arsiv["Toplam Gösterim"].sum()
            toplam_grp_a = round(df_arsiv["TR GRP"].sum(), 2)
            kapsanan_il_a, maks_erisim_a = hesapla_net_kapsama_metrikleri(df_arsiv, nufus_dict, TR_TOTAL_NUFUS)

            ak1.metric("📊 Toplam Gösterim", tr_tam_sayi(toplam_gos_a))
            ak2.metric("🇹🇷 Toplam TR GRP", tr_ondalik(toplam_grp_a, 2))
            ak3.metric("🌐 Maks. TR Erişimi", f"%{tr_ondalik(maks_erisim_a, 1)}")
            ak4.metric("📍 Kapsanan İl", f"{kapsanan_il_a} İl")

            rows_arsiv_html = "".join([
                f"<tr><td>{r['Yıl']}</td><td>{r['Dönem (Ay)']}</td><td>{r['Marka']}</td><td>{r['Kampanya Adı']}</td><td>{r['Mecra Adı']}</td><td>{r['Ünite']}</td><td>{r['İl']}</td><td>{r['Süre (Gün)']}</td><td>{r['Periyod']}</td><td>{tr_tam_sayi(r['Adet'])}</td><td>{tr_tam_sayi(r['Toplam Gösterim'])}</td><td>{tr_ondalik(r['Frekans'], 1)}</td><td>{tr_tam_sayi(r['Erişim (Kişi)'])}</td><td>{tr_tam_sayi(r['İl Nüfusu'])}</td><td>{tr_tam_sayi(r['TR Nüfusu'])}</td><td>%{tr_ondalik(r['TR Erişim %'], 2)}</td><td>{tr_ondalik(r['TR GRP'], 2)}</td></tr>"
                for _, r in df_arsiv.iterrows()
            ])
            
            table_arsiv_markup = f"""<div class="table-responsive-box"><table class="custom-ooh-table"><thead><tr><th>Yıl</th><th>Dönem</th><th>Marka</th><th>Kampanya</th><th>Mecra</th><th>Ünite</th><th>İl</th><th>Süre (Gün)</th><th>Periyod</th><th>Adet</th><th>Toplam Gösterim</th><th>Frekans</th><th>Erişim (Kişi)</th><th>İl Nüfusu</th><th>TR Nüfusu</th><th>TR Erişim %</th><th>TR GRP</th></tr></thead><tbody>{rows_html}</tbody></table></div>"""
            st.markdown(table_arsiv_markup, unsafe_allow_html=True)

            col_a1, col_a2, col_a3, col_a4, col_a5 = st.columns([1, 1.2, 1, 1.2, 1.2])
            with col_a1:
                if st.button("↩️ Son Satırı Sil", key="ars_del_last", use_container_width=True):
                    if st.session_state.arsiv_rows:
                        st.session_state.arsiv_rows.pop()
                        st.rerun()
            with col_a2:
                with st.popover("🗑️ Seçili Satırı Sil", use_container_width=True):
                    silinecek_idx = st.selectbox(
                        "Silinecek Satır No:",
                        range(len(st.session_state.arsiv_rows)),
                        key="ars_del_select",
                        format_func=lambda i: f"Satır {i+1}: {st.session_state.arsiv_rows[i]['Marka']} - {st.session_state.arsiv_rows[i]['Ünite']} ({st.session_state.arsiv_rows[i]['Mecra Adı']})"
                    )
                    if st.button("❌ Bu Satırı Sil", key="ars_del_btn", type="primary", use_container_width=True):
                        st.session_state.arsiv_rows.pop(silinecek_idx)
                        st.rerun()
            with col_a3:
                if st.button("🧹 Tümünü Temizle", key="ars_clear_all", use_container_width=True):
                    st.session_state.arsiv_rows = []
                    st.rerun()
            with col_a4:
                arsiv_html = generate_html_report(df_arsiv, "OOH Kampanya Arşiv & Yönetim Raporu", include_looker=False, is_arsiv=True)
                st.download_button(
                    label="📄 HTML Raporu Al",
                    data=arsiv_html,
                    file_name="OOH_Kampanya_Arsiv_Raporu.html",
                    mime="text/html",
                    use_container_width=True
                )
            with col_a5:
                arsiv_excel = generate_excel_report(df_arsiv, "OOH Kampanya Arşiv & Yönetim Raporu", looker_link=looker_url, is_arsiv=True)
                st.download_button(
                    label="📊 Excel Raporu Al",
                    data=arsiv_excel,
                    file_name="OOH_Kampanya_Arsiv_Raporu.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

# --- 8. KURUMSAL DİPNOT (FOOTER) ---
st.markdown("<div class='corporate-footer'>📌 CAFAS verileri dikkate alınarak geliştirilmiştir.</div>", unsafe_allow_html=True)
